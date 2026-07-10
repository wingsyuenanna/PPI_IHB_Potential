"""Export EU manufacturing facilities from the facility master and sync annual output into the heat demand workbook."""

from __future__ import annotations

import argparse
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[2]
DEFAULT_INPUT = PROJECT_ROOT / "Input" / "facility_master_2024_v6.csv"
DEFAULT_OUTPUT = PROJECT_ROOT / "heat_demand" / "facilities" / "facilities_2024_eu.csv"
DEFAULT_WORKBOOK = (
    PROJECT_ROOT / "heat_demand" / "pulp_paper_heat_demand_corrected.xlsx"
)
FACILITIES_SHEET = "Facilities"
ANNUAL_OUTPUT_COLUMN = 6  # column F: annual_output_t
DATA_START_ROW = 3
# The facility master's manufacturing sector spans every industrial subsector
# (food-beverage, textiles, glass, cement, iron-and-steel, pulp-and-paper, lime,
# chemicals, other-metals, petrochemical, aluminum). Power generation and
# mineral extraction are separate sectors and are excluded.
SECTOR = "manufacturing"

# EU28 (EU27 + United Kingdom). The facility master currently has no GBR
# pulp-and-paper sources, so the exported set matches EU27 in practice.
EU_ISO3_COUNTRIES = {
    "AUT", "BEL", "BGR", "HRV", "CYP", "CZE", "DNK", "EST", "FIN", "FRA",
    "DEU", "GRC", "HUN", "IRL", "ITA", "LVA", "LTU", "LUX", "MLT", "NLD",
    "POL", "PRT", "ROU", "SVK", "SVN", "ESP", "SWE", "GBR",
}

# Columns carried through from the facility master, in output order.
MASTER_COLUMNS = [
    "source_id",
    "source_name",
    "source_type",
    "iso3_country",
    "sector",
    "subsector",
    "lat",
    "lon",
    "capacity_units",
    "capacity",
    "capacity_factor",
    "co2e_100yr",
]


def export_facilities_2024(
    path: Path,
    year: int = 2024,
    sector: str = SECTOR,
    subsectors: list[str] | None = None,
    eu_only: bool = True,
) -> pd.DataFrame:
    """Return one row per EU manufacturing facility from the facility master.

    ``annual_output_t`` is derived from the updated production capacity as
    ``capacity * capacity_factor``. ``annual_capacity_t`` preserves the updated
    production capacity under the name used by downstream consumers. Pass
    ``subsectors`` to restrict the export to specific industries.
    """
    data = pd.read_csv(path)

    facilities = data[data["sector"] == sector].copy()
    if subsectors:
        facilities = facilities[facilities["subsector"].isin(subsectors)]
    if eu_only:
        facilities = facilities[facilities["iso3_country"].isin(EU_ISO3_COUNTRIES)]

    if facilities.empty:
        raise ValueError(
            f"No {sector} facilities found for the requested region in {path}"
        )

    facilities = facilities[MASTER_COLUMNS].copy()
    facilities["year"] = year
    facilities["annual_capacity_t"] = facilities["capacity"]
    facilities["annual_output_t"] = (
        facilities["capacity"] * facilities["capacity_factor"]
    )
    facilities = facilities.rename(
        columns={"co2e_100yr": "annual_emissions_co2e_t"}
    )

    return facilities.sort_values(["iso3_country", "source_name"]).reset_index(
        drop=True
    )


def sync_annual_output_to_workbook(
    workbook_path: Path,
    facilities: pd.DataFrame,
    sheet_name: str = FACILITIES_SHEET,
) -> tuple[int, int, list[int]]:
    """
    Write annual_output_t into the Facilities sheet by source_id.

    Returns (rows_updated, rows_missing_output, source_ids_not_in_workbook).
    """
    output_by_id = dict(
        zip(facilities["source_id"], facilities["annual_output_t"], strict=False)
    )
    workbook = load_workbook(workbook_path)
    worksheet = workbook[sheet_name]

    updated = 0
    missing_output: list[int] = []
    not_in_workbook: list[int] = []

    workbook_ids: set[int] = set()
    for row in range(DATA_START_ROW, worksheet.max_row + 1):
        source_id = worksheet.cell(row, 1).value
        if source_id is None or str(source_id).strip() == "":
            continue
        source_id = int(source_id)
        workbook_ids.add(source_id)

        if source_id not in output_by_id:
            continue

        annual_output = output_by_id[source_id]
        if pd.isna(annual_output):
            missing_output.append(source_id)
            continue

        worksheet.cell(row, ANNUAL_OUTPUT_COLUMN).value = float(annual_output)
        updated += 1

    for source_id in output_by_id:
        if source_id not in workbook_ids:
            not_in_workbook.append(source_id)

    workbook.save(workbook_path)
    return updated, len(missing_output), not_in_workbook


def main() -> None:
    parser = argparse.ArgumentParser(
        description=(
            "Export EU manufacturing facilities from the facility master and "
            "optionally sync annual_output_t into "
            "pulp_paper_heat_demand_corrected.xlsx."
        )
    )
    parser.add_argument(
        "--input",
        type=Path,
        default=DEFAULT_INPUT,
        help="Path to the facility master CSV.",
    )
    parser.add_argument(
        "--subsector",
        action="append",
        dest="subsectors",
        help=(
            "Restrict the export to one or more manufacturing subsectors "
            "(repeatable, e.g. --subsector cement --subsector glass). "
            "Defaults to all manufacturing subsectors."
        ),
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=DEFAULT_OUTPUT,
        help="Path to write the EU facilities CSV.",
    )
    parser.add_argument(
        "--workbook",
        type=Path,
        default=DEFAULT_WORKBOOK,
        help="Path to the formula-driven heat demand workbook.",
    )
    parser.add_argument(
        "--year",
        type=int,
        default=2024,
        help="Calendar year label for the exported facilities.",
    )
    parser.add_argument(
        "--no-sync-workbook",
        action="store_true",
        help="Export CSV only; do not update the heat demand workbook.",
    )
    args = parser.parse_args()

    facilities = export_facilities_2024(
        args.input, year=args.year, subsectors=args.subsectors
    )
    args.output.parent.mkdir(parents=True, exist_ok=True)
    facilities.to_csv(args.output, index=False)
    print(f"Wrote {len(facilities)} EU facilities to {args.output}")

    if not args.no_sync_workbook:
        if not args.workbook.exists():
            raise FileNotFoundError(f"Workbook not found: {args.workbook}")

        updated, missing, not_in_wb = sync_annual_output_to_workbook(
            args.workbook, facilities
        )
        print(
            f"Updated annual_output_t for {updated} facilities in {args.workbook.name}"
        )
        if missing:
            print(f"Skipped {missing} facilities with missing annual output")
        if not_in_wb:
            print(
                f"Note: {len(not_in_wb)} exported facilities are not in the workbook "
                f"(e.g. {not_in_wb[:5]})"
            )
        print("Open the workbook to view formula-driven heat demand results.")


if __name__ == "__main__":
    main()
