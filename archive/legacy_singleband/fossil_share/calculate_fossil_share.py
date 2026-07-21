"""Export fossil-share lookup table from the heat demand workbook."""

from __future__ import annotations

import argparse
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[2]
DEFAULT_WORKBOOK = (
    PROJECT_ROOT / "heat_demand" / "pulp_paper_heat_demand_corrected.xlsx"
)
DEFAULT_OUTPUT = PROJECT_ROOT / "heat_demand" / "fossil_share" / "fossil_share_lookup.csv"


def export_fossil_share_lookup(workbook_path: Path) -> pd.DataFrame:
    """
    Export country|NACE fossil shares from the workbook Fossil_Shares sheet.

    The workbook is the source of truth for fossil shares used in heat demand
    formulas. Eurostat 2023 data populates columns D–E in that sheet.
    """
    workbook = load_workbook(workbook_path, data_only=True)
    worksheet = workbook["Fossil_Shares"]

    records: list[dict] = []
    for row in range(3, worksheet.max_row + 1):
        country = worksheet.cell(row, 1).value
        if not country:
            continue

        total_tj = worksheet.cell(row, 4).value
        fossil_sum_tj = worksheet.cell(row, 5).value
        fossil_share = worksheet.cell(row, 6).value
        if fossil_share is None and total_tj not in (None, 0) and fossil_sum_tj is not None:
            fossil_share = fossil_sum_tj / total_tj

        country_str = str(country).strip()
        nace = worksheet.cell(row, 3).value
        lookup_key = f"{country_str}|{nace}"

        records.append(
            {
                "country": country_str,
                "sector": worksheet.cell(row, 2).value,
                "nace": nace,
                "total_tj": total_tj,
                "fossil_sum_tj": fossil_sum_tj,
                "fossil_share": fossil_share,
                "data_available": worksheet.cell(row, 7).value,
                "lookup_key": lookup_key,
            }
        )

    return pd.DataFrame(records)


def main() -> None:
    parser = argparse.ArgumentParser(
        description=(
            "Export fossil-share lookup values from pulp_paper_heat_demand_corrected.xlsx. "
            "Update Fossil_Shares in the workbook when Eurostat data changes."
        )
    )
    parser.add_argument(
        "--workbook",
        type=Path,
        default=DEFAULT_WORKBOOK,
        help="Path to the heat demand workbook.",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=DEFAULT_OUTPUT,
        help="Path to write the fossil share lookup CSV.",
    )
    args = parser.parse_args()

    result = export_fossil_share_lookup(args.workbook)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    result.to_csv(args.output, index=False)

    print(f"Wrote {len(result)} lookup rows to {args.output}")
    preview = result[result["data_available"] == "Yes"].head(10)
    print(
        preview[["country", "nace", "fossil_share"]].to_string(index=False)
    )


if __name__ == "__main__":
    main()
